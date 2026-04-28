from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from unittest.mock import AsyncMock, patch

import httpx
import openpyxl
import pytest

from brakerscalp.config import Settings
from brakerscalp.domain.models import DataHealth, Direction, MarketCandle, ScoreContribution, SetupType, SignalClass, SignalDecision, Timeframe, UniverseSymbol, Venue
from brakerscalp.services.api_service import build_api
from brakerscalp.signals.engine import ScreeningResult
from brakerscalp.universe import save_universe


@pytest.mark.asyncio
async def test_command_center_root_and_services_pages_render(repository, cache) -> None:
    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
    )
    app = build_api(repository, cache, settings, universe=[], adapters={})
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
        root = await client.get("/")
        services = await client.get("/services")

    assert root.status_code == 200
    assert "/screener" in root.text
    assert "/statistics" in root.text
    assert services.status_code == 200
    assert "PostgreSQL" in services.text


@pytest.mark.asyncio
async def test_statistics_page_and_threshold_route_render(repository, cache) -> None:
    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
    )
    app = build_api(repository, cache, settings, universe=[], adapters={})
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        statistics = await client.get("/statistics")
        export = await client.get("/statistics/export.xlsx")
        apply_threshold = await client.get("/settings/apply-threshold?value=74.5")
        apply_risk = await client.get("/settings/apply-risk?value=33.25")

    assert statistics.status_code == 200
    assert "Setup Statistics" in statistics.text
    assert "/statistics?range=week" in statistics.text
    assert "Export Excel" in statistics.text
    assert 'formaction="/statistics/export.xlsx"' in statistics.text
    assert export.status_code == 200
    assert export.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment; filename=" in export.headers["content-disposition"]
    assert apply_threshold.status_code == 303
    assert apply_threshold.headers["location"] == "/settings?threshold_saved=1"
    assert apply_risk.status_code == 303
    assert apply_risk.headers["location"] == "/settings?risk_saved=1"
    assert await cache.get_minimum_alert_confidence(65.0) == 74.5
    assert await cache.get_risk_usdt(25.0) == 33.25


@pytest.mark.asyncio
async def test_statistics_export_contains_trade_simulation_columns(repository, cache) -> None:
    detected_at = datetime.now(tz=timezone.utc) - timedelta(days=1)
    decision = SignalDecision(
        symbol="BTCUSDT",
        venue=Venue.BINANCE,
        timeframe=Timeframe.M15,
        setup=SetupType.BREAKOUT,
        direction=Direction.LONG,
        signal_class=SignalClass.ACTIONABLE,
        confidence=88.0,
        level_id="btc-breakout-level",
        alert_key="btc-breakout-test",
        detected_at=detected_at,
        entry_price=100.0,
        invalidation_price=95.0,
        targets=[105.0, 110.0],
        expected_rr=1.2,
        rationale=["Impulse confirmed", "Volume expansion"],
        why_not_higher=["Limited live history"],
        contributions=[ScoreContribution(group="level", score=20.0, max_score=25.0, reason="Strong level")],
        data_health=DataHealth(venue=Venue.BINANCE, symbol="BTCUSDT", is_fresh=True, freshness_ms=0),
        feature_snapshot={"atr_15m": 2.5},
        render_context={"trigger": "15m close above 100.0", "price_zone": "99.0 - 100.0"},
    )
    await repository.save_signal(decision)
    await repository.upsert_candles(
        [
            MarketCandle(
                symbol="BTCUSDT",
                venue=Venue.BINANCE,
                timeframe=Timeframe.M15,
                open_time=detected_at - timedelta(minutes=15),
                close_time=detected_at,
                open=99.0,
                high=101.0,
                low=98.5,
                close=100.5,
                volume=1000.0,
                quote_volume=100500.0,
                trade_count=10,
                taker_buy_volume=550.0,
                vwap=100.0,
            ),
            MarketCandle(
                symbol="BTCUSDT",
                venue=Venue.BINANCE,
                timeframe=Timeframe.M15,
                open_time=detected_at,
                close_time=detected_at + timedelta(minutes=15),
                open=100.5,
                high=106.0,
                low=100.0,
                close=105.5,
                volume=1300.0,
                quote_volume=137150.0,
                trade_count=12,
                taker_buy_volume=700.0,
                vwap=103.0,
            ),
        ]
    )

    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
    )
    app = build_api(repository, cache, settings, universe=[], adapters={})
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        export = await client.get(
            f"/statistics/export.xlsx?range=custom&start={(detected_at.date() - timedelta(days=1)).isoformat()}&end={detected_at.date().isoformat()}"
        )

    workbook = openpyxl.load_workbook(BytesIO(export.content))
    signals_sheet = workbook["signals"]
    headers = [cell.value for cell in next(signals_sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(signals_sheet.iter_rows(min_row=2, max_row=2))]

    assert "Trigger" in headers
    assert "Rationale" in headers
    assert "Entry price" in headers
    assert "TP1 price" in headers
    assert "TP2 price" in headers
    assert "SL price" in headers
    assert "Entry date" in headers
    assert "TP1 date" in headers
    assert "Final PnL %" in headers
    assert "Trade duration" in headers
    assert "BTCUSDT" in values


@pytest.mark.asyncio
async def test_statistics_export_uses_submitted_dates_even_when_range_is_not_custom(repository, cache) -> None:
    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
    )
    app = build_api(repository, cache, settings, universe=[], adapters={})
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        export = await client.get("/statistics/export.xlsx?range=day&start=2026-04-20&end=2026-04-22")

    workbook = openpyxl.load_workbook(BytesIO(export.content))
    summary_sheet = workbook["summary"]
    summary_rows = {summary_sheet[f"A{row}"].value: summary_sheet[f"B{row}"].value for row in range(2, 6)}

    assert summary_rows["Range"] == "custom"
    assert summary_rows["Start date"] == "2026-04-20"
    assert summary_rows["End date"] == "2026-04-22"


@pytest.mark.asyncio
async def test_statistics_page_persists_by_symbol_snapshot_in_database(repository, cache) -> None:
    detected_at = datetime.now(tz=timezone.utc) - timedelta(days=1)
    decision = SignalDecision(
        symbol="ETHUSDT",
        venue=Venue.BINANCE,
        timeframe=Timeframe.M15,
        setup=SetupType.BREAKOUT,
        direction=Direction.LONG,
        signal_class=SignalClass.ACTIONABLE,
        confidence=91.0,
        level_id="eth-breakout-level",
        alert_key="eth-breakout-test",
        detected_at=detected_at,
        entry_price=2000.0,
        invalidation_price=1975.0,
        targets=[2050.0, 2100.0],
        expected_rr=2.0,
        rationale=["Impulse confirmed"],
        why_not_higher=["Awaiting more samples"],
        contributions=[ScoreContribution(group="level", score=20.0, max_score=25.0, reason="Strong level")],
        data_health=DataHealth(venue=Venue.BINANCE, symbol="ETHUSDT", is_fresh=True, freshness_ms=0),
        feature_snapshot={"atr_15m": 12.0},
        render_context={"trigger": "15m close above 2000.0", "price_zone": "1995.0 - 2000.0"},
    )
    await repository.save_signal(decision)
    await repository.upsert_candles(
        [
            MarketCandle(
                symbol="ETHUSDT",
                venue=Venue.BINANCE,
                timeframe=Timeframe.M15,
                open_time=detected_at - timedelta(minutes=15),
                close_time=detected_at,
                open=1990.0,
                high=2003.0,
                low=1988.0,
                close=2001.0,
                volume=900.0,
                quote_volume=1800900.0,
                trade_count=8,
                taker_buy_volume=500.0,
                vwap=1997.0,
            ),
            MarketCandle(
                symbol="ETHUSDT",
                venue=Venue.BINANCE,
                timeframe=Timeframe.M15,
                open_time=detected_at,
                close_time=detected_at + timedelta(minutes=15),
                open=2001.0,
                high=2055.0,
                low=1999.0,
                close=2048.0,
                volume=1200.0,
                quote_volume=2457600.0,
                trade_count=13,
                taker_buy_volume=700.0,
                vwap=2024.5,
            ),
        ]
    )

    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
    )
    app = build_api(repository, cache, settings, universe=[], adapters={})
    transport = httpx.ASGITransport(app=app)
    start_value = (detected_at.date() - timedelta(days=1)).isoformat()
    end_value = detected_at.date().isoformat()
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        response = await client.get(f"/statistics?range=custom&start={start_value}&end={end_value}&q=ETH")

    start_at = datetime.combine(detected_at.date() - timedelta(days=1), datetime.min.time(), tzinfo=timezone.utc)
    end_at = datetime.combine(detected_at.date() + timedelta(days=1), datetime.min.time(), tzinfo=timezone.utc)
    persisted_rows = await repository.list_statistics_snapshot(start_at=start_at, end_at=end_at, symbol_query="ETH")

    assert response.status_code == 200
    assert "ETHUSDT" in response.text
    assert len(persisted_rows) == 1
    assert persisted_rows[0].symbol == "ETHUSDT"
    assert persisted_rows[0].success == 1
    assert persisted_rows[0].total == 1


@pytest.mark.asyncio
async def test_setups_page_gracefully_handles_missing_status_filter(repository, cache) -> None:
    detected_at = datetime.now(tz=timezone.utc) - timedelta(hours=6)
    decision = SignalDecision(
        symbol="SOLUSDT",
        venue=Venue.BINANCE,
        timeframe=Timeframe.M15,
        setup=SetupType.BREAKOUT,
        direction=Direction.LONG,
        signal_class=SignalClass.ACTIONABLE,
        confidence=87.0,
        level_id="sol-breakout-level",
        alert_key="sol-breakout-test",
        detected_at=detected_at,
        entry_price=150.0,
        invalidation_price=147.0,
        targets=[156.0, 162.0],
        expected_rr=2.0,
        rationale=["Momentum aligned"],
        why_not_higher=["Sample size is still limited"],
        contributions=[ScoreContribution(group="level", score=20.0, max_score=25.0, reason="Strong level")],
        data_health=DataHealth(venue=Venue.BINANCE, symbol="SOLUSDT", is_fresh=True, freshness_ms=0),
        feature_snapshot={"atr_15m": 1.8},
        render_context={"trigger": "15m close above 150.0", "price_zone": "149.0 - 150.0"},
    )
    await repository.save_signal(decision)
    await repository.upsert_candles(
        [
            MarketCandle(
                symbol="SOLUSDT",
                venue=Venue.BINANCE,
                timeframe=Timeframe.M15,
                open_time=detected_at - timedelta(minutes=15),
                close_time=detected_at,
                open=149.0,
                high=150.5,
                low=148.5,
                close=150.1,
                volume=1000.0,
                quote_volume=150100.0,
                trade_count=9,
                taker_buy_volume=560.0,
                vwap=149.6,
            ),
            MarketCandle(
                symbol="SOLUSDT",
                venue=Venue.BINANCE,
                timeframe=Timeframe.M15,
                open_time=detected_at,
                close_time=detected_at + timedelta(minutes=15),
                open=150.1,
                high=157.0,
                low=149.9,
                close=156.2,
                volume=1300.0,
                quote_volume=203060.0,
                trade_count=14,
                taker_buy_volume=780.0,
                vwap=153.15,
            ),
        ]
    )
    await cache.set_risk_usdt(25.0)

    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
    )
    app = build_api(repository, cache, settings, universe=[], adapters={})
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        response = await client.get("/setups?status=failed&min_confidence=")

    assert response.status_code == 200
    assert "SOLUSDT" in response.text
    assert "Qty" in response.text
    assert "156.0000 (50.00 USDT) / 162.0000 (100.00 USDT)" in response.text
    assert 'value="tp1"' in response.text
    assert 'value="loss"' not in response.text


@pytest.mark.asyncio
async def test_setups_page_uses_first_call_time_for_grouped_setup(repository, cache) -> None:
    first_detected_at = datetime.now(tz=timezone.utc) - timedelta(hours=6)
    second_detected_at = first_detected_at + timedelta(hours=2)
    base_kwargs = dict(
        symbol="BTCUSDT",
        venue=Venue.BINANCE,
        timeframe=Timeframe.M5,
        setup=SetupType.BREAKOUT,
        direction=Direction.LONG,
        signal_class=SignalClass.ACTIONABLE,
        confidence=91.0,
        level_id="btc-level-1",
        entry_price=100.0,
        invalidation_price=95.0,
        targets=[110.0, 120.0],
        expected_rr=2.0,
        rationale=["Momentum aligned"],
        why_not_higher=["Need more samples"],
        contributions=[ScoreContribution(group="level", score=20.0, max_score=25.0, reason="Strong level")],
        data_health=DataHealth(venue=Venue.BINANCE, symbol="BTCUSDT", is_fresh=True, freshness_ms=0),
        feature_snapshot={"atr_15m": 2.0},
        render_context={"trigger": "5m close above 100.0", "price_zone": "99.0 - 100.0"},
    )
    await repository.save_signal(
        SignalDecision(
            decision_id="btc-first-call",
            alert_key="btc-first-call",
            detected_at=first_detected_at,
            **base_kwargs,
        )
    )
    await repository.save_signal(
        SignalDecision(
            decision_id="btc-second-call",
            alert_key="btc-second-call",
            detected_at=second_detected_at,
            **base_kwargs,
        )
    )
    await repository.upsert_candles(
        [
            MarketCandle(
                symbol="BTCUSDT",
                venue=Venue.BINANCE,
                timeframe=Timeframe.M5,
                open_time=first_detected_at - timedelta(minutes=5),
                close_time=first_detected_at,
                open=99.0,
                high=100.5,
                low=98.8,
                close=100.1,
                volume=1000.0,
                quote_volume=100100.0,
                trade_count=10,
                taker_buy_volume=520.0,
                vwap=99.8,
            )
        ]
    )

    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
        timezone="Asia/Bangkok",
    )
    app = build_api(repository, cache, settings, universe=[], adapters={})
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        response = await client.get("/setups")

    assert response.status_code == 200
    expected_first_call = first_detected_at.astimezone(timezone(timedelta(hours=7))).strftime("%d.%m.%Y %H:%M:%S")
    unexpected_second_call = second_detected_at.astimezone(timezone(timedelta(hours=7))).strftime("%d.%m.%Y %H:%M:%S")
    assert expected_first_call in response.text
    assert unexpected_second_call not in response.text


@pytest.mark.asyncio
async def test_screener_page_supports_sorting_and_tooltips(repository, cache) -> None:
    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
    )
    fake_rows = [
        ScreeningResult(
            symbol="BTCUSDT",
            venue=Venue.BINANCE,
            setup=SetupType.BREAKOUT,
            status="monitor",
            confidence=87.0,
            direction=Direction.LONG,
            decision=None,
            last_price=100.0,
            level_id="btc-level",
            level_source="cascade-high",
            level_timeframe=Timeframe.H1,
            level_lower=99.0,
            level_upper=100.0,
            trend_bias=Direction.LONG,
            trend_score=0.8,
            coin_score=0.7,
            is_coin_in_play=True,
            atr_15m=1.2,
            volume_z_15m=2.2,
            volume_z_1h=1.4,
            range_expansion=1.6,
            quote_activity_ratio=1.8,
            squeeze_score=0.74,
            cascade_touches=4,
            consolidation_range_atr=1.1,
            breakout_distance_atr=0.18,
            body_ratio=0.6,
            follow_through_5m=True,
            book_imbalance=0.12,
            delta_ratio=0.18,
            cvd_slope=0.11,
            delta_divergence=False,
            freshness_ms=200,
            spread_ratio=1.0,
            notes=["Active breakout pressure"],
            updated_at=datetime.now(tz=timezone.utc),
        ),
        ScreeningResult(
            symbol="ETHUSDT",
            venue=Venue.BINANCE,
            setup=SetupType.BREAKOUT,
            status="watchlist",
            confidence=84.0,
            direction=Direction.SHORT,
            decision=None,
            last_price=90.0,
            level_id="eth-level",
            level_source="cascade-low",
            level_timeframe=Timeframe.H1,
            level_lower=89.0,
            level_upper=90.0,
            trend_bias=Direction.SHORT,
            trend_score=0.7,
            coin_score=0.6,
            is_coin_in_play=True,
            atr_15m=1.0,
            volume_z_15m=1.1,
            volume_z_1h=1.0,
            range_expansion=1.2,
            quote_activity_ratio=1.3,
            squeeze_score=0.81,
            cascade_touches=3,
            consolidation_range_atr=1.4,
            breakout_distance_atr=0.05,
            body_ratio=0.5,
            follow_through_5m=False,
            book_imbalance=-0.08,
            delta_ratio=-0.09,
            cvd_slope=-0.06,
            delta_divergence=False,
            freshness_ms=120,
            spread_ratio=1.1,
            notes=["Near level"],
            updated_at=datetime.now(tz=timezone.utc),
        ),
    ]
    with patch("brakerscalp.services.api_service.MarketInspector.screen_universe", new=AsyncMock(return_value=fake_rows)):
        app = build_api(repository, cache, settings, universe=[], adapters={})
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
            response = await client.get("/screener?sort_by=vol_z&sort_dir=asc")

    assert response.status_code == 200
    assert "Монета" in response.text
    assert "data-tooltip=\"HTF direction from the 1h/4h trend." in response.text
    assert "sort_by=symbol" in response.text
    assert response.text.index("ETHUSDT") < response.text.index("BTCUSDT")


@pytest.mark.asyncio
async def test_settings_strategy_routes_store_runtime_configuration(repository, cache) -> None:
    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
    )
    app = build_api(repository, cache, settings, universe=[], adapters={})
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        apply_response = await client.get(
            "/settings/apply-strategy"
            "?timeframe=5m"
            "&minimum_expected_rr=2.4"
            "&actionable_confidence_threshold=91"
            "&watchlist_confidence_threshold=84"
            "&volume_z_threshold=2.15"
            "&watchlist_volume_z_threshold=1.25"
            "&min_touches=4"
            "&squeeze_threshold=0.82"
            "&dist_to_level_atr=0.28"
            "&breakout_distance_atr=0.21"
            "&body_ratio_threshold=0.61"
            "&close_to_extreme_threshold=0.18"
            "&range_expansion_threshold=1.45"
            "&sl_multiplier=0.30"
            "&delta_ratio_threshold=0.16"
            "&watchlist_delta_ratio_threshold=0.05"
            "&cvd_slope_threshold=0.08"
            "&delta_divergence_threshold=0.07"
        )
        applied = await cache.get_strategy_config(default=settings.default_strategy_config())
        defaults_response = await client.get("/settings/strategy-defaults")
    restored = await cache.get_strategy_config(default=settings.default_strategy_config())

    assert apply_response.status_code == 303
    assert apply_response.headers["location"] == "/settings?strategy_saved=applied"
    assert applied["timeframe"] == "5m"
    assert applied["volume_z_threshold"] == 2.15
    assert applied["min_touches"] == 4
    assert applied["delta_ratio_threshold"] == 0.16
    assert defaults_response.status_code == 303
    assert defaults_response.headers["location"] == "/settings?strategy_saved=defaults"
    assert restored["timeframe"] == settings.default_strategy_config()["timeframe"]
    assert restored["volume_z_threshold"] == settings.default_strategy_config()["volume_z_threshold"]


@pytest.mark.asyncio
async def test_settings_universe_add_and_remove_updates_repository_and_cache(repository, cache, tmp_path: Path) -> None:
    universe_path = tmp_path / "universe.json"
    save_universe(universe_path, [UniverseSymbol(symbol="BTCUSDT", primary_venue=Venue.BINANCE)])
    await repository.replace_runtime_universe([UniverseSymbol(symbol="BTCUSDT", primary_venue=Venue.BINANCE)])
    await cache.store_universe([UniverseSymbol(symbol="BTCUSDT", primary_venue=Venue.BINANCE)])

    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
        universe_path=universe_path,
    )
    app = build_api(repository, cache, settings, universe=[UniverseSymbol(symbol="BTCUSDT", primary_venue=Venue.BINANCE)], adapters={Venue.BINANCE: object()})
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        add_response = await client.get("/settings/universe/add?symbol=ETH&venue=binance")
        remove_response = await client.get("/settings/universe/remove?symbol=BTCUSDT")

    runtime_universe = await repository.list_runtime_universe()
    cached_universe = await cache.get_universe_symbols()

    assert add_response.status_code == 303
    assert remove_response.status_code == 303
    assert {item.symbol for item in runtime_universe} == {"ETHUSDT"}
    assert {item.symbol for item in cached_universe} == {"ETHUSDT"}


@pytest.mark.asyncio
async def test_settings_page_prefers_persisted_runtime_universe_and_sorts_by_symbol(repository, cache, tmp_path: Path) -> None:
    universe_path = tmp_path / "universe.json"
    save_universe(universe_path, [UniverseSymbol(symbol="SOLUSDT", primary_venue=Venue.BINANCE)])
    await repository.replace_runtime_universe(
        [
            UniverseSymbol(symbol="ETHUSDT", primary_venue=Venue.BINANCE),
            UniverseSymbol(symbol="BTCUSDT", primary_venue=Venue.BYBIT),
        ]
    )
    await cache.store_universe([UniverseSymbol(symbol="SOLUSDT", primary_venue=Venue.BINANCE)])

    settings = Settings(
        _env_file=None,
        environment="test",
        bot_token="test-token",
        allowed_chat_ids=[1],
        alert_chat_ids=[1],
        database_url="sqlite+aiosqlite:///ignored.db",
        redis_url="redis://localhost:6379/0",
        universe_path=universe_path,
    )
    app = build_api(
        repository,
        cache,
        settings,
        universe=[UniverseSymbol(symbol="SOLUSDT", primary_venue=Venue.BINANCE)],
        adapters={Venue.BINANCE: object(), Venue.BYBIT: object()},
    )
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        response = await client.get("/settings")

    assert response.status_code == 200
    assert "BTCUSDT" in response.text
    assert "ETHUSDT" in response.text
    assert "SOLUSDT" not in response.text
    universe_section = response.text.split("Runtime Universe", 1)[1]
    assert universe_section.index("BTCUSDT") < universe_section.index("ETHUSDT")
